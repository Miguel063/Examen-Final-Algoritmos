"""
Desafío 2: Generación y modificación de un archivo Excel con openpyxl

Descripción:
Crear un script que cargue un archivo Excel existente, agregue algunos datos de forma programática,
y luego guarde el archivo modificado.

El código debe:
- Comprobar si el archivo Excel existe.
- Si el archivo no existe, debe crear uno nuevo y agregar un encabezado.
- Si el archivo ya existe, debe agregar una nueva fila con datos.
- Ajustar el tamaño de las columnas de acuerdo con el contenido.
- Mostrar mensajes de error adecuados si hay problemas al guardar el archivo o si el archivo tiene una extensión incorrecta.

Requisitos:
- Utilizar openpyxl para manejar el archivo Excel.
- Utilizar os para verificar la existencia del archivo.

"""

import openpyxl
import os
from openpyxl.utils import get_column_letter


def add_notas(nomb, nota, mate, sheet):
    sheet.append([nomb, nota, mate])
    
def update_excel(file_path):
    if not os.path.exists(file_path):
        print("El archivo no existe.")
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Nombre", "Nota", "Asignatura"])
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)
    

    
    print("Los datos fueron agregados correctamente al archivo Excel.")
    
        
    mas_notas = []
    while True:    
        print("\nIngrese El nombre del estudiante, con su nota y matria de la misma")
        nomb = input("Ingrese el Nombre del Estudiante: ")
        nota = input("Ingrese su nota: ")
        mate = input("Ingrese la materia: ")

        mas_notas.append([nomb, nota, mate])

        add_notas(nomb, nota, mate, sheet)

        continuar = input("¿Deseas agregar otra nota? (si/no): ").lower()
        if continuar != 'si':
            break
        
    try:
        wb.save(file_path)
        print("Se han guardado las notas")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
    
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    wb.close()
    if not file_path.endswith('.xlsx'):
        print("Error: El archivo debe tener la extensión '.xlsx'.")
        return
    


file_path = "notas.xlsx"
update_excel(file_path)
